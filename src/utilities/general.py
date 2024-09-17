import importlib.metadata
import io
import os
import platform
import subprocess
import openpyxl
import fitz
from docx import Document
from fastapi import HTTPException, UploadFile
from huggingface_hub import snapshot_download, hf_hub_download
from dotenv import load_dotenv


def check_gpu_linux():
    """
    Check if linux host has gpu support.
    :return: True if GPU is supported, False otherwise
    """
    try:
        # This command checks for CUDA installation, indicating GPU support.
        result = subprocess.run(["nvcc", "--version"], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        return result.returncode == 0
    except FileNotFoundError:
        print("CUDA not found. GPU support may not be enabled.")
        return False
    except Exception as e:
        print(f"Error checking GPU on Linux: {e}")
        return False


# Load environment variables
load_dotenv(os.getenv("ENV", "config/.env-dev"))

# Define constants
PLATFORM = platform.system() if platform.system() != 'Darwin' else "MacOS"
BASE_MODEL_DIR = os.getenv("BASE_MODEL_DIR", "/opt/models")
EFS_MODEL_DIR = os.getenv("EFS_MODEL_DIR", "/efs/models/microsoft/Phi-3-vision-128k-instruct-onnx")
GPU_SUPPORTED = check_gpu_linux()
VISION_MODEL_DIR = f"{EFS_MODEL_DIR}-{'cuda/cuda-int4-rtn-block-32' if GPU_SUPPORTED else 'cpu/cpu-int4-rtn-block-32-acc-level-4'}"
CONTEXT_WINDOW = int(os.getenv("CONTEXT_WINDOW", "131072"))


def run_command(command, cwd=None, check=True):
    """
    Run command and return output.
    :param command: Command to run as list of strings
    :param cwd: Current working directory
    :param check: Flag to raise an error if the command fails
    """
    try:
        print(f"Running command in directory {cwd}: {' '.join(command)}")
        subprocess.run(command, cwd=cwd, check=check)
    except subprocess.CalledProcessError as e:
        print(f"Command failed with error: {e}")
        raise


def download_vision_model():
    """Downloads the Vision model if not already present in the EFS directory."""
    repo_id = "microsoft/Phi-3-vision-128k-instruct-onnx-cpu" if not GPU_SUPPORTED else "microsoft/Phi-3-vision-128k-instruct-onnx-cuda"
    expected_files = [
        "genai_config.json",
        "phi-3-v-128k-instruct-text-embedding.onnx",
        "phi-3-v-128k-instruct-text-embedding.onnx.data",
        "phi-3-v-128k-instruct-text.onnx",
        "phi-3-v-128k-instruct-text.onnx.data",
        "phi-3-v-128k-instruct-vision.onnx",
        "phi-3-v-128k-instruct-vision.onnx.data",
        "processor_config.json",
        "special_tokens_map.json",
        "tokenizer.json",
        "tokenizer_config.json",
    ]

    if os.path.exists(VISION_MODEL_DIR):
        print(f"Model directory exists at {VISION_MODEL_DIR}. Skipping download.")
        return

    print("Downloading Vision model...")
    partial_dir = '/'.join(VISION_MODEL_DIR.split("/")[:-1])
    os.makedirs(partial_dir, exist_ok=True)

    try:
        print(f"Downloading the repository {repo_id} to {partial_dir}...")
        snapshot_download(repo_id=repo_id, local_dir=partial_dir)
    except Exception as e:
        print(f"Error downloading repository Phi-3-vision: {e}")
        return

    # Check for missing files
    downloaded_files = os.listdir(VISION_MODEL_DIR)
    missing_files = [f for f in expected_files if f not in downloaded_files]

    if missing_files:
        print(f"Missing files: {missing_files}. Re-downloading missing files...")
        for file in missing_files:
            try:
                hf_hub_download(repo_id=repo_id, filename=file, local_dir=partial_dir)
            except Exception as e:
                print(f"Error downloading missing file {file}: {e}")


def check_package(package_name):
    """
    Checks if a package is installed via pip.
    :param package_name: Name of the package to check
    :return: True if the package is installed, False otherwise
    """
    try:
        importlib.metadata.version(package_name)
        return True
    except importlib.metadata.PackageNotFoundError:
        return False


async def extract_text_from_pdf(file: UploadFile):
    try:
        file_data = await file.read()  # Read the file content into memory
        doc = fitz.open(stream=file_data, filetype="pdf")
        text = "".join(page.get_text() for page in doc)
        return text
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process PDF: {e}")


async def extract_text_from_docx(file: UploadFile):
    try:
        file_data = await file.read()  # Read the file content into memory
        file_stream = io.BytesIO(file_data)  # Create a file-like object in memory
        doc = Document(file_stream)
        text = "\n".join(paragraph.text for paragraph in doc.paragraphs)  # Extract text from each paragraph
        return text
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process DOCX: {e}")


async def extract_text_from_excel(file: UploadFile):
    try:
        file_data = await file.read()  # Read the file content into memory
        file_stream = io.BytesIO(file_data)  # Create a file-like object in memory
        workbook = openpyxl.load_workbook(file_stream, data_only=True)
        text = ""
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            text += f"Sheet: {sheet}\n"  # Add sheet name
            for row in ws.iter_rows(values_only=True):
                text += " ".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
        return text
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process XLSX: {e}")


async def extract_text_from_txt(file: UploadFile):
    try:
        content = await file.read()
        text_content = content.decode("utf-8")  # Decode bytes to string
        return text_content
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process TXT: {e}")
